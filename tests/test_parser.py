"""
Tests for core/parser.py — multi-sheet and legacy format parsing.
"""

import datetime
import pytest
import openpyxl
import pandas as pd
from pathlib import Path

from core.parser import (
    parse_log,
    build_medication_periods,
    get_active_medications,
    _is_date,
    _normalise_med_name,
    _parse_meals_sheet,
    _parse_bac_sheet,
    _parse_med_sheet,
    MEDICATION_ALIASES,
)


# ---------------------------------------------------------------------------
# Helpers — create in-memory Excel workbooks for testing
# ---------------------------------------------------------------------------
def _make_multi_sheet_workbook(tmp_path: Path, *, meals=None, bac=None, meds=None):
    """Create a multi-sheet .xlsx file and return its path."""
    wb = openpyxl.Workbook()

    # Meals sheet
    ws_meals = wb.active
    ws_meals.title = "Meals"
    ws_meals.append(["Date", "Meal", "Time", "Product", "Measure", "g",
                     "kcal", "Protein", "Fat", "Sat.Fat", "Carbs", "Sugars", "Fibre"])
    if meals:
        for row in meals:
            ws_meals.append(row)
    else:
        # Default test data: 1 day, 1 meal, 2 ingredients
        ws_meals.append([datetime.date(2025, 4, 1)])
        ws_meals.append([None, "Breakfast", datetime.time(8, 0)])
        ws_meals.append([None, None, None, "Oatmeal", "bowl", 250,
                         None, None, None, None, 45, 5, 8])
        ws_meals.append([None, None, None, "Banana", "1 medium", 120,
                         None, None, None, None, 27, 14, 3])

    # Bac Log sheet
    ws_bac = wb.create_sheet("Bac Log")
    ws_bac.append(["Date", "Time", "BAC‰", "Comment"])
    if bac:
        for row in bac:
            ws_bac.append(row)
    else:
        ws_bac.append([datetime.date(2025, 4, 1)])
        ws_bac.append([None, datetime.time(9, 0), 0.3])
        ws_bac.append([None, datetime.time(13, 0), 1.5, "After lunch"])

    # Medications sheet
    ws_med = wb.create_sheet("Medications")
    ws_med.append(["Date", "Medication", "Action", "Comment"])
    if meds:
        for row in meds:
            ws_med.append(row)
    else:
        ws_med.append([datetime.date(2025, 3, 28)])
        ws_med.append([None, "Fluconazole", "Start", "200mg daily"])

    path = tmp_path / "test_log.xlsx"
    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# Unit tests — small helpers
# ---------------------------------------------------------------------------
class TestIsDate:
    def test_datetime_is_date(self):
        assert _is_date(datetime.datetime(2025, 1, 1)) is True

    def test_date_is_not_datetime(self):
        # _is_date only matches datetime.datetime, not datetime.date
        assert _is_date(datetime.date(2025, 1, 1)) is False

    def test_string_is_not_date(self):
        assert _is_date("2025-01-01") is False

    def test_none_is_not_date(self):
        assert _is_date(None) is False

    def test_number_is_not_date(self):
        assert _is_date(42) is False


class TestNormaliseMedName:
    def test_known_alias(self):
        assert _normalise_med_name("activated charcol") == "Activated Charcoal"

    def test_unknown_med_title_cased(self):
        assert _normalise_med_name("some new drug") == "Some New Drug"

    def test_empty_string(self):
        assert _normalise_med_name("") == ""


class TestBuildMedicationPeriods:
    def test_start_and_stop(self):
        events = [
            {"date": datetime.date(2025, 1, 1), "medication": "DrugA", "action": "start"},
            {"date": datetime.date(2025, 1, 10), "medication": "DrugA", "action": "stop"},
        ]
        periods = build_medication_periods(events)
        assert "DrugA" in periods
        assert len(periods["DrugA"]) == 1
        assert periods["DrugA"][0]["start"] == datetime.date(2025, 1, 1)
        assert periods["DrugA"][0]["stop"] == datetime.date(2025, 1, 10)

    def test_open_ended(self):
        events = [
            {"date": datetime.date(2025, 1, 1), "medication": "DrugB", "action": "start"},
        ]
        periods = build_medication_periods(events)
        assert periods["DrugB"][0]["stop"] is None

    def test_multiple_periods(self):
        events = [
            {"date": datetime.date(2025, 1, 1), "medication": "DrugA", "action": "start"},
            {"date": datetime.date(2025, 1, 5), "medication": "DrugA", "action": "stop"},
            {"date": datetime.date(2025, 2, 1), "medication": "DrugA", "action": "start"},
            {"date": datetime.date(2025, 2, 10), "medication": "DrugA", "action": "stop"},
        ]
        periods = build_medication_periods(events)
        assert len(periods["DrugA"]) == 2

    def test_empty_events(self):
        assert build_medication_periods([]) == {}


class TestGetActiveMedications:
    def test_active_during_period(self):
        periods = {"DrugA": [{"start": datetime.date(2025, 1, 1), "stop": datetime.date(2025, 1, 10)}]}
        assert get_active_medications(datetime.date(2025, 1, 5), periods) == ["DrugA"]

    def test_not_active_outside_period(self):
        periods = {"DrugA": [{"start": datetime.date(2025, 1, 1), "stop": datetime.date(2025, 1, 10)}]}
        assert get_active_medications(datetime.date(2025, 2, 1), periods) == []

    def test_open_ended_period(self):
        periods = {"DrugA": [{"start": datetime.date(2025, 1, 1), "stop": None}]}
        assert get_active_medications(datetime.date(2099, 12, 31), periods) == ["DrugA"]

    def test_multiple_active(self):
        periods = {
            "DrugA": [{"start": datetime.date(2025, 1, 1), "stop": None}],
            "DrugB": [{"start": datetime.date(2025, 1, 5), "stop": None}],
        }
        result = get_active_medications(datetime.date(2025, 1, 10), periods)
        assert result == ["DrugA", "DrugB"]


# ---------------------------------------------------------------------------
# Integration tests — multi-sheet parsing
# ---------------------------------------------------------------------------
class TestParseLogMultiSheet:
    def test_basic_parse(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path)
        meals_df, bac_df, med_periods = parse_log(path)

        assert not meals_df.empty
        assert not bac_df.empty
        assert len(meals_df) == 2  # Oatmeal, Banana
        assert len(bac_df) == 2   # two readings
        assert "Fluconazole" in med_periods

    def test_ingredients_parsed_correctly(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path)
        meals_df, _, _ = parse_log(path)

        ingredients = set(meals_df["ingredient"].tolist())
        assert "Oatmeal" in ingredients
        assert "Banana" in ingredients

    def test_bac_sorted_by_datetime(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path)
        _, bac_df, _ = parse_log(path)

        datetimes = bac_df["bac_datetime"].tolist()
        assert datetimes == sorted(datetimes)

    def test_active_medications_column(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path)
        _, bac_df, _ = parse_log(path)

        # Fluconazole started March 28, readings April 1 → should be active
        assert all("Fluconazole" in str(m) for m in bac_df["active_medications"])

    def test_episode_threshold(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path)
        _, bac_df, _ = parse_log(path)

        # 0.3 → not episode, 1.5 → not episode (threshold is 2.0)
        assert bac_df["episode"].sum() == 0

    def test_comment_field(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path)
        _, bac_df, _ = parse_log(path)

        comments = bac_df["comment"].tolist()
        assert "After lunch" in comments

    def test_carbs_parsed(self, tmp_path):
        path = _make_multi_sheet_workbook(tmp_path)
        meals_df, _, _ = parse_log(path)

        oat = meals_df[meals_df["ingredient"] == "Oatmeal"].iloc[0]
        assert oat["carbs_g"] == 45
        assert oat["sugars_g"] == 5

    def test_split_compounds(self, tmp_path):
        meals = [
            [datetime.date(2025, 4, 1)],
            [None, "Lunch", datetime.time(12, 0)],
            [None, None, None, "Pasta & Tomato Sauce", "plate", 300,
             None, None, None, None, 60, 8, 4],
        ]
        path = _make_multi_sheet_workbook(tmp_path, meals=meals)

        # With split
        meals_df, _, _ = parse_log(path, split_compounds=True)
        ingredients = set(meals_df["ingredient"].tolist())
        assert "Pasta" in ingredients
        assert "Tomato Sauce" in ingredients

        # Without split
        meals_df2, _, _ = parse_log(path, split_compounds=False)
        assert "Pasta & Tomato Sauce" in meals_df2["ingredient"].tolist()

    def test_medication_period_stop(self, tmp_path):
        meds = [
            [datetime.date(2025, 3, 28)],
            [None, "Fluconazole", "Start"],
            [datetime.date(2025, 4, 5)],
            [None, "Fluconazole", "Stop"],
        ]
        path = _make_multi_sheet_workbook(tmp_path, meds=meds)
        _, _, med_periods = parse_log(path)

        assert med_periods["Fluconazole"][0]["stop"] == datetime.date(2025, 4, 5)

    def test_med_name_normalisation(self, tmp_path):
        meds = [
            [datetime.date(2025, 3, 1)],
            [None, "activated charcol", "Start"],
        ]
        path = _make_multi_sheet_workbook(tmp_path, meds=meds)
        _, _, med_periods = parse_log(path)

        assert "Activated Charcoal" in med_periods

    def test_empty_bac_sheet(self, tmp_path):
        # Pass bac rows that only have a date header but no readings
        bac = [[datetime.date(2025, 4, 1)]]
        path = _make_multi_sheet_workbook(tmp_path, bac=bac)
        meals_df, bac_df, _ = parse_log(path)
        assert bac_df.empty

    def test_multiple_days(self, tmp_path):
        meals = [
            [datetime.date(2025, 4, 1)],
            [None, "Breakfast", datetime.time(8, 0)],
            [None, None, None, "Rice", "cup", 200, None, None, None, None, 50, 0, 1],
            [datetime.date(2025, 4, 2)],
            [None, "Lunch", datetime.time(12, 0)],
            [None, None, None, "Bread", "slice", 50, None, None, None, None, 25, 2, 1],
        ]
        bac = [
            [datetime.date(2025, 4, 1)],
            [None, datetime.time(10, 0), 0.5],
            [datetime.date(2025, 4, 2)],
            [None, datetime.time(14, 0), 1.8],
        ]
        path = _make_multi_sheet_workbook(tmp_path, meals=meals, bac=bac)
        meals_df, bac_df, _ = parse_log(path)

        assert len(meals_df) == 2
        assert len(bac_df) == 2
        dates = bac_df["date"].dt.date.unique()
        assert len(dates) == 2


# ---------------------------------------------------------------------------
# Integration tests — the example workbook shipped with the app
# ---------------------------------------------------------------------------
_REPO_ROOT = Path(__file__).resolve().parents[1]
EXAMPLE_LOG = _REPO_ROOT / "example" / "example_log.xlsx"


class TestParseExampleLog:
    """Parse `example/example_log.xlsx`, the workbook offered on the main page.

    The synthetic fixtures above are clean by construction; this file is not.
    Its Meals sheet interleaves daily-total and meal-header rows (which carry
    quantities and calories but no product name) with real ingredient rows,
    leaves nutrient cells blank, and pads every sheet with ~1300 empty rows.

    Counts below describe the workbook as committed — update them if the
    example data is regenerated.
    """

    @pytest.fixture(scope="class")
    def parsed(self):
        return parse_log(EXAMPLE_LOG)

    def test_file_is_committed(self):
        assert EXAMPLE_LOG.exists(), f"missing example workbook: {EXAMPLE_LOG}"

    def test_expected_columns(self, parsed):
        meals_df, bac_df, med_periods = parsed
        assert "ingredient" in meals_df.columns
        assert "bac_datetime" in bac_df.columns
        assert "active_medications" in bac_df.columns
        assert isinstance(med_periods, dict)

    def test_row_counts(self, parsed):
        # 15 ingredient rows in the sheet, 16 after the one compound dish splits.
        meals_df, bac_df, _ = parsed
        assert len(meals_df) == 16
        assert len(bac_df) == 5
        assert meals_df["ingredient"].nunique() == 9

    def test_aggregate_rows_are_not_ingredients(self, parsed):
        meals_df, _, _ = parsed
        assert meals_df["ingredient"].notna().all()
        assert (meals_df["ingredient"].str.strip() != "").all()
        assert set(meals_df["meal"].unique()) == {"Breakfast", "Lunch", "Dinner"}

    def test_padding_rows_are_skipped(self, parsed):
        meals_df, bac_df, _ = parsed
        assert meals_df["meal_datetime"].notna().all()
        assert bac_df["bac_datetime"].notna().all()
        assert bac_df["promille"].notna().all()

    def test_covers_three_days(self, parsed):
        _, bac_df, _ = parsed
        assert bac_df["date"].dt.date.nunique() == 3

    def test_bac_sorted_by_datetime(self, parsed):
        _, bac_df, _ = parsed
        datetimes = bac_df["bac_datetime"].tolist()
        assert datetimes == sorted(datetimes)

    def test_medication_periods(self, parsed):
        _, _, med_periods = parsed
        assert set(med_periods) == {"Activated Charcoal", "Rifaximin"}
        charcoal = med_periods["Activated Charcoal"][0]
        assert charcoal["start"] == datetime.date(2026, 3, 22)
        assert charcoal["stop"] == datetime.date(2026, 3, 23)

    def test_active_medications_follow_periods(self, parsed):
        """Charcoal stops on the 23rd, so the 24th should list Rifaximin only."""
        _, bac_df, _ = parsed
        per_date = bac_df.groupby(bac_df["date"].dt.date)["active_medications"].unique()
        assert list(per_date[datetime.date(2026, 3, 23)]) == [
            "Activated Charcoal, Rifaximin"
        ]
        assert list(per_date[datetime.date(2026, 3, 24)]) == ["Rifaximin"]

    def test_comment_preserved(self, parsed):
        _, bac_df, _ = parsed
        assert "problem talking" in bac_df["comment"].tolist()

    def test_no_episodes_below_threshold(self, parsed):
        _, bac_df, _ = parsed
        assert bac_df["promille"].max() == 0.8
        assert bac_df["episode"].sum() == 0

    def test_compound_dish_is_split(self, parsed):
        """'Broccoli & chicken soup' is the only compound row in the workbook."""
        meals_df, _, _ = parsed
        ingredients = set(meals_df["ingredient"])
        assert {"Broccoli", "Chicken"} <= ingredients
        assert "Broccoli & chicken soup" not in ingredients

    def test_compound_dish_kept_whole_when_splitting_disabled(self):
        meals_df, bac_df, _ = parse_log(EXAMPLE_LOG, split_compounds=False)
        ingredients = set(meals_df["ingredient"])
        assert "Broccoli & chicken soup" in ingredients
        assert not {"Broccoli", "Chicken"} & ingredients
        assert len(meals_df) == 15
        assert len(bac_df) == 5


# ---------------------------------------------------------------------------
# Integration tests — private legacy log (skipped unless present)
# ---------------------------------------------------------------------------
LEGACY_LOG = _REPO_ROOT / "data" / "jo_log_old.xlsx"


class TestRealDataFiles:
    @pytest.mark.skipif(
        not LEGACY_LOG.exists(),
        reason="Legacy real data file not available",
    )
    def test_parse_real_legacy(self):
        meals_df, bac_df, _ = parse_log(LEGACY_LOG)
        assert not meals_df.empty
        assert not bac_df.empty
