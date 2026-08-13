"""Generate a synthetic year-scale ABS log workbook (research R8).

Produces the multi-sheet format `core/parser.py` reads ("Meals", "Bac Log",
"Medications"), reproducing the structural quirks that make the real files awkward:

  - date rows interleaved with data rows (the parser's carry-forward state machine),
  - blank padding rows between days,
  - aggregate/total rows that are not data and must be skipped,
  - partially filled nutrient columns (some ingredient rows lack carbs/sugars),
  - compound dishes joined with "&" so compound splitting is exercised,
  - protein ingredients so `exclude_proteins` filtering is exercised.

It contains NO real patient data (Principle IV) and is generated on demand rather
than committed. Numbers are seeded so a given `--months` is reproducible.

Usage:
    uv run python tests/fixtures/generate_year_log.py --months 12 --out /tmp/year_log.xlsx
"""

from __future__ import annotations

import argparse
import datetime as dt
import random
from pathlib import Path

import openpyxl

# Ingredient pools. Compound entries use "&" so the splitter has work to do;
# protein entries let exclude_proteins filtering be measured.
_CARB_INGREDIENTS = [
    "White Rice", "Brown Bread", "Pasta", "Potato", "Banana", "Apple Juice",
    "Oats", "Honey", "Sugar", "Mango", "Grapes", "Cola", "Beer-Free Malt",
    "Sourdough", "Couscous", "Sweet Potato", "Dates", "Raisins",
]
_PROTEIN_INGREDIENTS = [
    "Chicken", "Beef", "Salmon", "Eggs", "Tuna", "Pork", "Turkey", "Cod",
]
_COMPOUND_DISHES = [
    "Lentil & Kale soup",
    "Chicken & Rice bowl",
    "Beef & Potato stew",
    "Tomato & Cucumber salad",
    "Oats & Banana mix",
]
_MEALS = [("Breakfast", 8), ("Lunch", 13), ("Snack", 16), ("Dinner", 19)]
_MEDS = ["Rifaximin", "Vancomycin", "Fluconazole"]


def _gen_meals(ws, days: list[dt.date], rng: random.Random) -> None:
    ws.append(
        ["Date", "Meal", "Time", "Product", "Measure", "g", "kcal",
         "Protein", "Fat", "Sat", "Carbs", "Sugars", "Fibre"]
    )
    for day in days:
        ws.append([dt.datetime(day.year, day.month, day.day)])
        for meal_name, base_hour in _MEALS:
            meal_time = dt.time(base_hour, rng.choice([0, 15, 30, 45]))
            row = [None, meal_name, meal_time]
            ws.append(row)
            n_ingredients = rng.randint(2, 4)
            for _ in range(n_ingredients):
                roll = rng.random()
                if roll < 0.20:
                    product = rng.choice(_COMPOUND_DISHES)
                elif roll < 0.50:
                    product = rng.choice(_PROTEIN_INGREDIENTS)
                else:
                    product = rng.choice(_CARB_INGREDIENTS)
                grams = rng.randint(40, 300)
                # Partially filled nutrients: sometimes carbs/sugars are blank.
                if rng.random() < 0.15:
                    carbs = sugars = None
                else:
                    carbs = round(grams * rng.uniform(0.1, 0.7), 1)
                    sugars = round(carbs * rng.uniform(0.1, 0.6), 1)
                ws.append(
                    [None, None, None, product, "g", grams, None,
                     None, None, None, carbs, sugars, None]
                )
            # Aggregate/total row that is not data — parser must skip it
            # (no date, no meal label, no product name).
            if rng.random() < 0.10:
                ws.append([None, None, None, None, "Total", None, None])
        # Blank padding row between days.
        ws.append([None])


def _gen_bac(ws, days: list[dt.date], rng: random.Random) -> None:
    ws.append(["Date", "Time", "BAC", "Comment"])
    for day in days:
        ws.append([dt.datetime(day.year, day.month, day.day)])
        n_readings = rng.randint(4, 7)
        for _ in range(n_readings):
            hour = rng.randint(7, 23)
            reading_time = dt.time(hour, rng.choice([0, 20, 40]))
            # Skew mostly low with occasional episodes (>= 2.0 permille).
            if rng.random() < 0.12:
                bac = round(rng.uniform(2.0, 4.5), 2)
                comment = "felt unwell"
            else:
                bac = round(rng.uniform(0.0, 1.8), 2)
                comment = None
            ws.append([None, reading_time, bac, comment])
        if rng.random() < 0.10:
            ws.append([None])  # blank padding


def _gen_meds(ws, days: list[dt.date], rng: random.Random) -> None:
    ws.append(["Date", "Medication", "Action", "Comment"])
    if not days:
        return
    span = (days[-1] - days[0]).days or 1
    # The parser reads events from rows WITHOUT a date in column 0, carrying the
    # date forward from the most recent date row. So each event is a date row
    # followed by a separate medication row — matching the real workbook shape.
    events: list[tuple[dt.date, str, str]] = []
    for med in _MEDS:
        for _ in range(rng.randint(1, 2)):
            start_offset = rng.randint(0, max(0, span - 10))
            stop_offset = min(span, start_offset + rng.randint(7, 40))
            start = days[0] + dt.timedelta(days=start_offset)
            stop = days[0] + dt.timedelta(days=stop_offset)
            events.append((start, med, "start"))
            events.append((stop, med, "stop"))
    for day, med, action in sorted(events, key=lambda e: e[0]):
        ws.append([dt.datetime(day.year, day.month, day.day)])
        ws.append([None, med, action])


def generate(months: int, out: Path, seed: int = 1234) -> Path:
    """Write a synthetic workbook covering `months` of daily entries."""
    rng = random.Random(seed)
    start = dt.date(2024, 1, 1)
    n_days = int(round(months * 30.4))
    days = [start + dt.timedelta(days=i) for i in range(n_days)]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    _gen_meals(wb.create_sheet("Meals"), days, rng)
    _gen_bac(wb.create_sheet("Bac Log"), days, rng)
    _gen_meds(wb.create_sheet("Medications"), days, rng)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description="Generate a synthetic year-scale ABS log.")
    ap.add_argument("--months", type=float, default=12, help="months of daily entries")
    ap.add_argument("--out", type=Path, default=Path("/tmp/year_log.xlsx"))
    ap.add_argument("--seed", type=int, default=1234)
    args = ap.parse_args()
    path = generate(args.months, args.out, args.seed)
    print(f"wrote {path} ({args.months} months)")


if __name__ == "__main__":
    main()
