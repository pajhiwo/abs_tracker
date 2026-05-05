"""
Excel log parser — reads the ABS diet log into clean DataFrames.

Supports two formats:
  1. Multi-sheet (jo_log_new.xlsx): Sheet "Meals", "Bac Log", "Medications"
  2. Legacy single-sheet (jo_log.xlsx): all data in one sheet
"""

import datetime
import warnings
from pathlib import Path

import openpyxl
import pandas as pd

warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# ---------------------------------------------------------------------------
# Column indices (0-based) — legacy single-sheet format (jo_log.xlsx)
# ---------------------------------------------------------------------------
C_DATE = 0
C_MEAL = 1
C_MEAL_TIME = 2
C_PRODUCT = 3
C_MEASURE = 4
C_GRAMS = 5
C_CALORIES = 6
C_PROTEIN = 7
C_FAT = 8
C_SAT_FAT = 9
C_CARBS = 10
C_SUGARS = 11
C_FIBRE = 12
C_BAC_TIME = 13
C_BAC_VAL = 14
C_EPISODE = 15
C_MEDICATION = 16
C_COMMENT = 17

MEAL_LABELS = {"Breakfast", "Snack", "Lunch", "Dinner"}
HEADER_STRINGS = {"Date", "Meal", "Espisode", "Episode"}

# Known medication spelling corrections (lowercase) → canonical name.
MEDICATION_ALIASES = {
    "activated charcol": "Activated Charcoal",
    "charcol": "Activated Charcoal",
    "charcoal": "Activated Charcoal",
    "vancomicin": "Vancomycin",
}

# Default BAC threshold (‰) above which a reading is considered an "episode".
EPISODE_THRESHOLD = 2.0


# ---------------------------------------------------------------------------
# Multi-sheet format detection
# ---------------------------------------------------------------------------
_MULTI_SHEET_NAMES = {"Meals", "Bac Log", "Medications"}


def _is_multi_sheet(path: str | Path) -> bool:
    """Return True if the workbook has the expected multi-sheet layout."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    names = set(wb.sheetnames)
    wb.close()
    return _MULTI_SHEET_NAMES.issubset(names)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _is_date(val) -> bool:
    return isinstance(val, datetime.datetime)


def _normalise_med_name(raw: str) -> str:
    """Normalise a medication name via aliases or title-case."""
    lower = raw.strip().lower()
    for alias, canonical in MEDICATION_ALIASES.items():
        if alias in lower:
            return canonical
    return raw.strip().title()


# ---------------------------------------------------------------------------
# Multi-sheet parsers
# ---------------------------------------------------------------------------
def _parse_meals_sheet(
    ws,
    *,
    split_compounds: bool = True,
) -> list[dict]:
    """Parse the Meals worksheet (date-row → meal-row → ingredient-rows)."""
    rows_out: list[dict] = []
    current_date = None
    current_meal = None
    current_meal_time = None
    current_meal_dt = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Date(0) Meal(1) Time(2) Product(3) Measure(4) g(5)
        #          kcal(6) Protein(7) Fat(8) Sat(9) Carbs(10) Sugars(11) Fibre(12)
        val_date = row[0]
        val_meal = row[1] if len(row) > 1 else None
        val_time = row[2] if len(row) > 2 else None
        val_product = row[3] if len(row) > 3 else None
        val_grams = row[5] if len(row) > 5 else None
        val_carbs = row[10] if len(row) > 10 else None
        val_sugars = row[11] if len(row) > 11 else None

        # Date row
        if _is_date(val_date):
            current_date = val_date.date()
            current_meal = None
            current_meal_time = None
            current_meal_dt = None
            continue

        if current_date is None:
            continue

        # Meal row (has meal label + optional time, no product)
        if val_meal and str(val_meal) in MEAL_LABELS:
            current_meal = str(val_meal)
            if isinstance(val_time, datetime.time):
                current_meal_time = val_time
                current_meal_dt = datetime.datetime.combine(current_date, val_time)
            else:
                current_meal_time = None
                current_meal_dt = None
            continue

        # Ingredient row (has product name)
        if val_product and str(val_product).strip():
            raw_name = str(val_product).strip()

            if split_compounds and "&" in raw_name:
                parts = [p.strip() for p in raw_name.split("&")]
                last_part = parts[-1]
                suffix = ""
                last_words = last_part.rsplit(None, 1)
                if len(last_words) == 2:
                    suffix = last_words[1].lower()
                    if suffix in ("soup", "stew", "salad", "bowl", "mix"):
                        parts[-1] = last_words[0]
                    else:
                        suffix = ""
                ingredient_names = [p.strip().title() for p in parts if p.strip()]
            else:
                ingredient_names = [raw_name]

            grams = float(val_grams) if val_grams is not None else None
            carbs = float(val_carbs) if val_carbs is not None else None
            sugars = float(val_sugars) if val_sugars is not None else None

            for ingredient in ingredient_names:
                rows_out.append(
                    {
                        "date": current_date,
                        "meal": current_meal,
                        "meal_time": current_meal_time,
                        "meal_datetime": current_meal_dt,
                        "ingredient": ingredient,
                        "quantity_g": grams,
                        "carbs_g": carbs,
                        "sugars_g": sugars,
                    }
                )

    return rows_out


def _parse_bac_sheet(ws) -> list[dict]:
    """Parse the Bac Log worksheet (date-row → reading-rows)."""
    rows_out: list[dict] = []
    current_date = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Date(0) Time(1) BAC‰(2) Comment(3)
        val_date = row[0]
        val_time = row[1] if len(row) > 1 else None
        val_bac = row[2] if len(row) > 2 else None
        val_comment = row[3] if len(row) > 3 else None

        if _is_date(val_date):
            current_date = val_date.date()
            continue

        if current_date is None:
            continue

        if val_bac is not None:
            bac_dt = (
                datetime.datetime.combine(current_date, val_time)
                if isinstance(val_time, datetime.time)
                else None
            )
            comment = (
                str(val_comment).strip()
                if val_comment is not None and str(val_comment).strip()
                else None
            )
            rows_out.append(
                {
                    "date": current_date,
                    "bac_time": (
                        val_time if isinstance(val_time, datetime.time) else None
                    ),
                    "bac_datetime": bac_dt,
                    "promille": float(val_bac),
                    "comment": comment,
                }
            )

    return rows_out


def _parse_med_sheet(ws) -> list[dict]:
    """Parse the Medications worksheet (date-row → event-rows)."""
    events: list[dict] = []
    current_date = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Columns: Date(0) Medication(1) Action(2) Comment(3)
        val_date = row[0]
        val_med = row[1] if len(row) > 1 else None
        val_action = row[2] if len(row) > 2 else None

        if _is_date(val_date):
            current_date = val_date.date()
            continue

        if current_date is None:
            continue

        if val_med and val_action:
            action = str(val_action).strip().lower()
            if action in ("start", "stop"):
                med_name = _normalise_med_name(str(val_med))
                events.append(
                    {"date": current_date, "medication": med_name, "action": action}
                )

    return sorted(events, key=lambda e: e["date"])


# ---------------------------------------------------------------------------
# build_medication_periods
# ---------------------------------------------------------------------------
def build_medication_periods(events: list[dict]) -> dict[str, list[dict]]:
    """
    Convert start/stop events into date ranges per medication.
    Returns: {medication_name: [{"start": date, "stop": date|None}, ...]}
    """
    periods: dict[str, list[dict]] = {}
    open_starts: dict[str, datetime.date] = {}

    for event in events:
        med = event["medication"]
        date = event["date"]

        if event["action"] == "start":
            open_starts[med] = date

        elif event["action"] == "stop":
            if med in open_starts:
                periods.setdefault(med, []).append(
                    {
                        "start": open_starts.pop(med),
                        "stop": date,
                    }
                )

    # Any medication still open at end of data → stop=None (ongoing)
    for med, start in open_starts.items():
        periods.setdefault(med, []).append({"start": start, "stop": None})

    return periods


# ---------------------------------------------------------------------------
# get_active_medications
# ---------------------------------------------------------------------------
def get_active_medications(
    date: datetime.date,
    periods: dict[str, list[dict]],
) -> list[str]:
    """Return list of medications active on a given date."""
    active = []
    for med, ranges in periods.items():
        for r in ranges:
            stop = r["stop"] or datetime.date(9999, 12, 31)
            if r["start"] <= date <= stop:
                active.append(med)
                break
    return sorted(active)


# ---------------------------------------------------------------------------
# parse_log — main entry point (auto-detects format)
# ---------------------------------------------------------------------------
def parse_log(
    path: str | Path,
    *,
    split_compounds: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """
    Parse the Excel log into two clean DataFrames plus medication periods.

    Auto-detects format:
      - Multi-sheet: sheets "Meals", "Bac Log", "Medications"
      - Legacy: single sheet with all data

    Returns
    -------
    meals_df    : one row per ingredient
    bac_df      : one row per BAC reading, with active_medications column
    med_periods : {medication: [{start, stop}, ...]}
    """
    if _is_multi_sheet(path):
        return _parse_log_multi(path, split_compounds=split_compounds)
    return _parse_log_legacy(path, split_compounds=split_compounds)


# ---------------------------------------------------------------------------
# Multi-sheet parse_log
# ---------------------------------------------------------------------------
def _parse_log_multi(
    path: str | Path,
    *,
    split_compounds: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """Parse multi-sheet format (jo_log_new.xlsx)."""
    wb = openpyxl.load_workbook(path, data_only=True)

    # Medications first (needed for active_medications on BAC rows)
    med_events = _parse_med_sheet(wb["Medications"])
    med_periods = build_medication_periods(med_events)

    # Meals
    meals_rows = _parse_meals_sheet(wb["Meals"], split_compounds=split_compounds)
    meals_df = pd.DataFrame(meals_rows)

    # BAC
    bac_rows = _parse_bac_sheet(wb["Bac Log"])
    # Enrich BAC rows with active_medications
    for row in bac_rows:
        active_meds = get_active_medications(row["date"], med_periods)
        row["active_medications"] = ", ".join(active_meds) if active_meds else "none"
    bac_df = pd.DataFrame(bac_rows)

    wb.close()

    if not meals_df.empty:
        meals_df["date"] = pd.to_datetime(meals_df["date"])
    if not bac_df.empty:
        bac_df["date"] = pd.to_datetime(bac_df["date"])
        bac_df["bac_datetime"] = pd.to_datetime(bac_df["bac_datetime"])
        bac_df["episode"] = bac_df["promille"] >= EPISODE_THRESHOLD
        bac_df = bac_df.sort_values("bac_datetime").reset_index(drop=True)

    return meals_df, bac_df, med_periods


# ---------------------------------------------------------------------------
# Legacy: parse_medication_events (single-sheet format)
# ---------------------------------------------------------------------------
def parse_medication_events(raw: pd.DataFrame) -> list[dict]:
    """
    Scan date rows in legacy single-sheet format for medication entries.
    Returns: [{"date": date, "medication": str, "action": "start"|"stop"}, ...]
    """
    events = []
    date_rows = raw[raw[C_DATE].apply(lambda v: isinstance(v, datetime.datetime))]

    for _, row in date_rows.iterrows():
        med_raw = row[C_MEDICATION]
        if pd.isna(med_raw):
            continue
        date = row[C_DATE].date()
        for part in str(med_raw).split(","):
            part = part.strip()
            part_lower = part.lower()
            action = None
            if "start" in part_lower:
                action = "start"
            elif "stop" in part_lower:
                action = "stop"
            else:
                continue

            med_name = _normalise_med_name(
                part_lower.replace("start", "")
                .replace("stop", "")
                .strip()
                .strip("-–— ")
            )
            if med_name and action:
                events.append({"date": date, "medication": med_name, "action": action})

    return sorted(events, key=lambda e: e["date"])


# ---------------------------------------------------------------------------
# Legacy single-sheet parse_log
# ---------------------------------------------------------------------------
def _parse_log_legacy(
    path: str | Path,
    *,
    split_compounds: bool = True,
) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """Parse legacy single-sheet format (jo_log.xlsx)."""
    raw = pd.read_excel(path, sheet_name=0, header=None)
    raw = raw[~raw[C_DATE].astype(str).isin(HEADER_STRINGS)].reset_index(drop=True)

    # Build medication periods first
    events = parse_medication_events(raw)
    med_periods = build_medication_periods(events)

    meals_rows = []
    bac_rows = []

    current_date = None
    current_meal = None
    current_meal_time = None
    current_meal_dt = None

    for _, row in raw.iterrows():
        has_date = pd.notna(row[C_DATE]) and isinstance(row[C_DATE], datetime.datetime)
        has_meal = pd.notna(row[C_MEAL]) and str(row[C_MEAL]) in MEAL_LABELS
        has_product = pd.notna(row[C_PRODUCT])
        has_bac = pd.notna(row[C_BAC_VAL])

        if has_date:
            current_date = row[C_DATE].date()
            current_meal = None
            current_meal_time = None
            current_meal_dt = None

        if has_meal and current_date is not None:
            current_meal = str(row[C_MEAL])
            raw_time = row[C_MEAL_TIME]
            if isinstance(raw_time, datetime.time):
                current_meal_time = raw_time
                current_meal_dt = datetime.datetime.combine(current_date, raw_time)
            else:
                current_meal_time = None
                current_meal_dt = None

        if has_bac and current_date is not None:
            bac_time = row[C_BAC_TIME]
            bac_dt = (
                datetime.datetime.combine(current_date, bac_time)
                if isinstance(bac_time, datetime.time)
                else None
            )
            episode_raw = row[C_EPISODE]
            is_episode = (
                pd.notna(episode_raw) and str(episode_raw).strip().lower() == "yes"
            )
            active_meds = get_active_medications(current_date, med_periods)
            bac_rows.append(
                {
                    "date": current_date,
                    "bac_time": bac_time,
                    "bac_datetime": bac_dt,
                    "promille": float(row[C_BAC_VAL]),
                    "episode": is_episode,
                    "active_medications": (
                        ", ".join(active_meds) if active_meds else "none"
                    ),
                    "comment": row[C_COMMENT] if pd.notna(row[C_COMMENT]) else None,
                }
            )

        if has_product and not has_date and not has_meal and current_date is not None:
            grams = row[C_GRAMS] if pd.notna(row[C_GRAMS]) else None
            raw_name = str(row[C_PRODUCT]).strip()

            # Split compound ingredients on "&" (e.g. "Zuccini & Chicken & Spinach soup")
            # The trailing word after the last ingredient (e.g. "soup") is stripped from each part
            if split_compounds and "&" in raw_name:
                parts = [p.strip() for p in raw_name.split("&")]
                # Check if the last part ends with a shared suffix like "soup"
                last_part = parts[-1]
                suffix = ""
                last_words = last_part.rsplit(None, 1)
                if len(last_words) == 2:
                    suffix = last_words[1].lower()
                    # Only treat as suffix if it's a known dish type
                    if suffix in ("soup", "stew", "salad", "bowl", "mix"):
                        parts[-1] = last_words[0]  # remove suffix from last part
                    else:
                        suffix = ""
                ingredient_names = [p.strip().title() for p in parts if p.strip()]
            else:
                ingredient_names = [raw_name]

            for ingredient in ingredient_names:
                carbs = row[C_CARBS] if pd.notna(row[C_CARBS]) else None
                sugars = row[C_SUGARS] if pd.notna(row[C_SUGARS]) else None
                meals_rows.append(
                    {
                        "date": current_date,
                        "meal": current_meal,
                        "meal_time": current_meal_time,
                        "meal_datetime": current_meal_dt,
                        "ingredient": ingredient,
                        "quantity_g": float(grams) if grams is not None else None,
                        "carbs_g": float(carbs) if carbs is not None else None,
                        "sugars_g": float(sugars) if sugars is not None else None,
                    }
                )

    meals_df = pd.DataFrame(meals_rows)
    bac_df = pd.DataFrame(bac_rows)

    if not meals_df.empty:
        meals_df["date"] = pd.to_datetime(meals_df["date"])
    if not bac_df.empty:
        bac_df["date"] = pd.to_datetime(bac_df["date"])
        bac_df["bac_datetime"] = pd.to_datetime(bac_df["bac_datetime"])
        bac_df = bac_df.sort_values("bac_datetime").reset_index(drop=True)

    return meals_df, bac_df, med_periods
